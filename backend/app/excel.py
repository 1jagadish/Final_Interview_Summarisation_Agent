import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")

os.makedirs(DATA_DIR, exist_ok=True)

FILE = os.path.join(DATA_DIR, "results.xlsx")


# ✅ SAFE CONVERSION FUNCTION (CRITICAL FIX)
def safe_list(value):
    if isinstance(value, list):
        return value
    elif isinstance(value, str):
        return [value]
    else:
        return []


def save_to_excel(meta, result):

    # ✅ SAFELY HANDLE DATA TYPES
    strengths_list = safe_list(result.get("strengths", []))
    improvements_list = safe_list(result.get("improvements", []))

    strengths = "\n".join(strengths_list)
    improvements = "\n".join(improvements_list)

    data = {
        "Candidate Name": meta.get("name", ""),
        "Role": meta.get("role", ""),
        "Date": meta.get("date", ""),
        "Summary": result.get("summary", ""),
        "Strengths": strengths,
        "Improvements": improvements,
        "Recommendation": result.get("recommendation", "Unknown"),
        "Analyzed At": datetime.now().strftime("%Y-%m-%d %H:%M")
    }

    df_new = pd.DataFrame([data])

    # ✅ Append or create safely
    try:
        if os.path.exists(FILE):
            df = pd.read_excel(FILE, engine="openpyxl")
            df = pd.concat([df, df_new], ignore_index=True)
        else:
            df = df_new
    except Exception as e:
        print("🔥 Excel read error:", e)
        df = df_new

    df.to_excel(FILE, index=False, engine="openpyxl")

    # ---------------------------------------
    # 🔥 FORMAT EXCEL (PROPER UI)
    # ---------------------------------------
    wb = load_workbook(FILE)
    ws = wb.active

    # ✅ Column widths
    widths = {
        "A": 22,
        "B": 28,
        "C": 15,
        "D": 60,
        "E": 50,
        "F": 50,
        "G": 20,
        "H": 22
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # ✅ Wrap text + align top
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ✅ Header styling
    for cell in ws[1]:
        cell.font = Font(bold=True)

    wb.save(FILE)